"""必须下单清单 Excel 导出。"""

from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from src.i18n import priority_label, t

MUST_ORDER_PRIORITIES = ("紧急", "高")
PRIORITY_SORT = {"紧急": 0, "高": 1, "中": 2, "低": 3, "充足": 4}

SHIPPING_OPTIONS = {
    "en": ["International Express", "Slow / Sea Freight", "Domestic", "In-store / Local"],
    "zh": ["国际快递", "海运/慢货", "国内物流", "自采/到店"],
}


def build_must_order_df(df: pd.DataFrame, lang: str) -> pd.DataFrame:
    """筛选必须立即下单的商品（紧急/高优先级，或库存≤0）。"""
    if df.empty:
        return pd.DataFrame()

    mask = (df["建议进货量"] > 0) & (
        df["优先级"].isin(MUST_ORDER_PRIORITIES) | (df["当前库存"] <= 0)
    )
    subset = df.loc[mask].copy()
    if subset.empty:
        return pd.DataFrame()

    subset["_sort"] = subset["优先级"].map(PRIORITY_SORT).fillna(9)
    subset = subset.sort_values(["_sort", "建议进货量"], ascending=[True, False]).drop(
        columns="_sort"
    )

    return pd.DataFrame(
        {
            t("mo_col_product", lang): subset["商品名称"],
            t("mo_col_sku", lang): subset["SKU_ID"],
            t("mo_col_stock", lang): subset["当前库存"],
            t("mo_col_reorder", lang): subset["建议进货量"],
            t("mo_col_priority", lang): subset["优先级"].map(
                lambda p: priority_label(p, lang)
            ),
            t("mo_col_lead_time", lang): "",
            t("mo_col_shipping", lang): "",
            t("mo_col_notes", lang): "",
        }
    ).reset_index(drop=True)


def _guide_rows(lang: str) -> list[list[str]]:
    if lang == "zh":
        return [
            [t("mo_guide_title", lang), ""],
            ["", ""],
            [t("mo_guide_shipping_header", lang), t("mo_guide_shipping_desc", lang)],
            ["国际快递", "DHL / FedEx 等，到货快但成本高，适合紧急补货"],
            ["海运/慢货", "韩日代购、海运拼箱等，到货慢（2-4周+），需提前下单"],
            ["国内物流", "国内仓发货，一般 3-7 天"],
            ["自采/到店", "自行采购或供应商送货上门"],
            ["", ""],
            [t("mo_guide_lead_header", lang), t("mo_guide_lead_desc", lang)],
            ["", "在「到货周期(天)」列填写预计到货天数，慢货务必标注"],
            ["", ""],
            [t("mo_guide_tip", lang), ""],
        ]
    return [
        [t("mo_guide_title", lang), ""],
        ["", ""],
        [t("mo_guide_shipping_header", lang), t("mo_guide_shipping_desc", lang)],
        ["International Express", "DHL / FedEx — fast, higher cost, for urgent items"],
        ["Slow / Sea Freight", "Sea freight / proxy buying — slow (2-4+ weeks), order early"],
        ["Domestic", "Domestic warehouse — typically 3-7 days"],
        ["In-store / Local", "Self-purchase or supplier delivery"],
        ["", ""],
        [t("mo_guide_lead_header", lang), t("mo_guide_lead_desc", lang)],
        ["", "Fill estimated lead time (days) in the Lead Time column"],
        ["", ""],
        [t("mo_guide_tip", lang), ""],
    ]


def _style_workbook(wb, lang: str, data_rows: int) -> None:
    ws = wb[t("mo_sheet_name", lang)]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    urgent_fill = PatternFill("solid", fgColor="FADBD8")
    high_fill = PatternFill("solid", fgColor="FDEBD0")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    priority_col_idx = 5
    for row_idx in range(2, data_rows + 2):
        pri_cell = ws.cell(row=row_idx, column=priority_col_idx)
        pri_val = str(pri_cell.value or "")
        urgent_label = priority_label("紧急", lang)
        high_label = priority_label("高", lang)
        row_fill = None
        if pri_val == urgent_label:
            row_fill = urgent_fill
        elif pri_val == high_label:
            row_fill = high_fill
        if row_fill:
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).fill = row_fill

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 24
    ws.freeze_panes = "A2"

    shipping_col = "G"
    options = ",".join(SHIPPING_OPTIONS.get(lang, SHIPPING_OPTIONS["en"]))
    dv = DataValidation(
        type="list",
        formula1=f'"{options}"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = t("mo_shipping_validation", lang)
    ws.add_data_validation(dv)
    if data_rows > 0:
        dv.add(f"{shipping_col}2:{shipping_col}{data_rows + 1}")

    guide = wb[t("mo_guide_sheet", lang)]
    guide.column_dimensions["A"].width = 22
    guide.column_dimensions["B"].width = 55
    guide["A1"].font = Font(bold=True, size=12)


def build_must_order_excel_bytes(df: pd.DataFrame, lang: str) -> bytes:
    order_df = build_must_order_df(df, lang)
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        if order_df.empty:
            pd.DataFrame({t("mo_empty", lang): []}).to_excel(
                writer, index=False, sheet_name=t("mo_sheet_name", lang),
            )
        else:
            order_df.to_excel(writer, index=False, sheet_name=t("mo_sheet_name", lang))

        guide_df = pd.DataFrame(_guide_rows(lang), columns=["A", "B"])
        guide_df.to_excel(
            writer, index=False, header=False, sheet_name=t("mo_guide_sheet", lang),
        )

    buffer.seek(0)
    wb = load_workbook(buffer)
    if not order_df.empty:
        _style_workbook(wb, lang, len(order_df))
    else:
        _style_workbook(wb, lang, 0)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def save_must_order_excel(
    df: pd.DataFrame,
    lang: str,
    output_dir: str | Path = "output",
) -> Path | None:
    order_df = build_must_order_df(df, lang)
    if order_df.empty:
        return None

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = t("mo_filename", lang, ts=timestamp)
    path = output_path / filename
    path.write_bytes(build_must_order_excel_bytes(df, lang))
    return path
